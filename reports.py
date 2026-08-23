"""
reports.py
Generates downloadable attendance reports in PDF and Excel formats
(document requirement 1.9.3 "Easy reporting features" / 4.8.4).
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.graphics.barcode import code128
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def build_excel_report(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Reg Number", "Full Name", "Course Code", "Course Name", "Lecturer", "Session Date", "Status", "Timestamp"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["reg_number"], r["full_name"], r["course_code"], r["course_name"],
            r.get("lecturer_name") or "—", r["session_date"], r["status"], r["timestamp"],
        ])

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(12, length + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf_report(rows, start_date=None, end_date=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), title="Attendance Report")
    styles = getSampleStyleSheet()
    elements = []

    title = "Attendance Report"
    elements.append(Paragraph(title, styles["Title"]))
    period = f"Period: {start_date or 'earliest'} to {end_date or 'latest'}"
    elements.append(Paragraph(period, styles["Normal"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 0.25 * inch))

    data = [["Reg Number", "Full Name", "Course", "Lecturer", "Session Date", "Status", "Timestamp"]]
    for r in rows:
        data.append([
            r["reg_number"], r["full_name"], f"{r['course_code']}", r.get("lecturer_name") or "—",
            r["session_date"], r["status"].upper(), r["timestamp"],
        ])

    if len(data) == 1:
        data.append(["-", "No records found for the selected filters", "-", "-", "-", "-", "-"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def build_id_cards_pdf(students, institution_name="Institution"):
    """
    Prints one wallet-sized ID card per student (3 columns x 4 rows per A4
    page), each with a real, scannable Code128 barcode — for students who
    don't have a physical card yet, or as a backup to the camera-scan flow.
    """
    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4

    cols, rows = 3, 4
    margin = 0.4 * inch
    card_w = (page_w - 2 * margin) / cols
    card_h = (page_h - 2 * margin) / rows

    for i, s in enumerate(students):
        pos = i % (cols * rows)
        if i > 0 and pos == 0:
            c.showPage()
        col = pos % cols
        row = pos // cols
        x = margin + col * card_w
        y = page_h - margin - (row + 1) * card_h

        pad = 8
        c.setStrokeColor(colors.HexColor("#1F2937"))
        c.roundRect(x + 4, y + 4, card_w - 8, card_h - 8, 6, stroke=1, fill=0)

        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(colors.HexColor("#1F2937"))
        c.drawCentredString(x + card_w / 2, y + card_h - pad - 8, institution_name.upper())

        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(x + card_w / 2, y + card_h - pad - 24, s["full_name"][:28])

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#555555"))
        c.drawCentredString(x + card_w / 2, y + card_h - pad - 36, s["reg_number"])
        if s.get("department"):
            c.drawCentredString(x + card_w / 2, y + card_h - pad - 47, s["department"][:30])

        barcode = code128.Code128(s["barcode_value"], barHeight=0.35 * inch, barWidth=0.9)
        bw = barcode.width
        bx = x + (card_w - bw) / 2
        by = y + 14
        barcode.drawOn(c, bx, by)

        c.setFont("Helvetica", 6)
        c.setFillColor(colors.HexColor("#888888"))
        c.drawCentredString(x + card_w / 2, y + 6, s["barcode_value"])

    c.save()
    buf.seek(0)
    return buf
